import requests
import time
import openpyxl

headers = {
    'content-type': 'application/x-www-form-urlencoded',
    'cookie': 'your cookies',
    'origin': 'https://live.bilibili.com',
    'referer': 'https://live.bilibili.com/12371301?visit_id=1xu2wxrhs928',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36'
}
data = {
    'roomid': '12371301',
    'csrf_token': 'a6d3068b4f42cabe0d3a6ae96124b2b1',
    'csrf': 'a6d3068b4f42cabe0d3a6ae96124b2b1',
    'visit_id': ''
}
url = 'https://api.live.bilibili.com/xlive/web-room/v1/dM/gethistory'

saves = []   # [uid, nickname, text]
wb = openpyxl.load_workbook('danmu.xlsx')   # 打开工作簿
sh = wb['Sheet1']                # 选取表单
iterator = list(sh.rows)
if len(list(sh.rows)) > 10:
    iterator = list(sh.rows)[-10:]
for item in iterator:
    saves.append([item[1].value, item[2].value, item[3].value, item[0].value])  # 获取数据库中最近10个弹幕，防止重复获取

def fun():
    time.sleep(2)
    response = requests.post(url=url, data=data, headers=headers)
    dic_data = response.json()
    for item in dic_data['data']['room']:
        tmp = [item['uid'], item['nickname'], item['text'], item['timeline']]
        if tmp not in saves:
            saves.append(tmp)
            _row = len(list(sh.rows)) + 1  # 获取行数
            sh.cell(row=_row, column=1, value=tmp[3])
            sh.cell(row=_row, column=2, value=tmp[0])
            sh.cell(row=_row, column=3, value=tmp[1])
            sh.cell(row=_row, column=4, value=tmp[2])

            try:
                wb.save('danmu.xlsx')
            except Exception as err:
                print(err)
                print('请关闭Excel !!!')
            try:
                ol = len(tmp[1].encode('GBK')) - len(tmp[1])
                print('{:<11}{:<23}{:<{out_len}}:   {}'.format(tmp[0], tmp[3], tmp[1], tmp[2], out_len=20-ol))
            except Exception as err:
                print('{:<11}{:<23}{:<{out_len}}:   {}'.format(tmp[0], tmp[3], tmp[1], tmp[2], out_len=20))
                print(err)

if __name__ == '__main__':
    while True:
        fun()



